import datetime
import shutil
from unittest.mock import ANY, MagicMock, Mock, patch

import openpyxl
import pytest
import pytz
from django.core.files import File
from freezegun import freeze_time

from saleor.core import JobStatus
from saleor.csv import ExportEvents, FileTypes
from saleor.csv.models import ExportEvent, ExportFile
from saleor.csv.utils.export import (
    create_csv_file_and_save_in_export_file,
    export_products,
    get_filename,
    get_product_queryset,
    on_task_failure,
    on_task_success,
    save_csv_file_in_export_file,
    update_export_file_when_task_finished,
)


@patch("saleor.csv.utils.export.send_export_failed_info")
def test_on_task_failure(send_export_failed_info_mock, export_file):
    exc = Exception("Test")
    task_id = "task_id"
    args = [export_file.pk, {"all": ""}]
    kwargs = {}
    info_type = "Test error"
    info = Mock(type=info_type)

    assert export_file.status == JobStatus.PENDING
    assert export_file.created_at
    previous_updated_at = export_file.updated_at

    on_task_failure(None, exc, task_id, args, kwargs, info)

    export_file.refresh_from_db()
    assert export_file.status == JobStatus.FAILED
    assert export_file.created_at
    assert export_file.updated_at != previous_updated_at
    export_failed_event = ExportEvent.objects.get(
        export_file=export_file,
        user=export_file.created_by,
        type=ExportEvents.EXPORT_FAILED,
    )
    assert export_failed_event.parameters == {
        "message": str(exc),
        "error_type": info_type,
    }

    send_export_failed_info_mock.called_once_with(export_file, "export_failed")


def test_on_task_success(export_file):
    task_id = "task_id"
    args = [export_file.pk, {"filter": {}}]
    kwargs = {}

    assert export_file.status == JobStatus.PENDING
    assert export_file.created_at
    previous_updated_at = export_file.updated_at

    on_task_success(None, None, task_id, args, kwargs)

    export_file.refresh_from_db()
    assert export_file.status == JobStatus.SUCCESS
    assert export_file.created_at
    assert export_file.updated_at != previous_updated_at
    assert ExportEvent.objects.filter(
        export_file=export_file,
        user=export_file.created_by,
        type=ExportEvents.EXPORT_SUCCESS,
    )


def test_update_export_file_when_task_finished(export_file):
    with freeze_time(datetime.datetime.now()) as frozen_datetime:
        previous_updated_at = export_file.updated_at
        update_export_file_when_task_finished(export_file, JobStatus.FAILED)

        export_file.refresh_from_db()
        assert export_file.updated_at == pytz.utc.localize(frozen_datetime())
        assert export_file.updated_at != previous_updated_at


@pytest.mark.parametrize(
    "scope, filet_type",
    [
        ({"filter": {"is_published": True}}, FileTypes.CSV),
        ({"all": ""}, FileTypes.XLSX),
    ],
)
@patch("saleor.csv.utils.export.send_email_with_link_to_download_csv")
@patch("saleor.csv.utils.export.save_csv_file_in_export_file")
def test_export_products(
    save_csv_file_in_export_file_mock,
    send_email_mock,
    product_list,
    export_file,
    scope,
    filet_type,
):
    export_info = {"fields": [], "warehouses": [], "attributes": []}
    export_products(export_file.id, scope, export_info, filet_type)

    save_csv_file_in_export_file_mock.called_once_with(export_file, ANY)
    send_email_mock.called_once_with(export_file)


@patch("saleor.csv.utils.export.send_email_with_link_to_download_csv")
@patch("saleor.csv.utils.export.save_csv_file_in_export_file")
def test_export_products_ids(
    save_csv_file_in_export_file_mock, send_email_mock, product_list, export_file
):
    pks = [product.pk for product in product_list[:2]]
    export_info = {"fields": [], "warehouses": [], "attributes": []}
    file_type = FileTypes.CSV

    assert export_file.status == JobStatus.PENDING
    assert not export_file.content_file

    export_products(export_file.id, {"ids": pks}, export_info, file_type)

    save_csv_file_in_export_file_mock.called_once_with(export_file, ANY)
    send_email_mock.called_once_with(export_file)


def test_get_filename_csv():
    with freeze_time("2000-02-09"):
        file_name = get_filename("test", FileTypes.CSV)

        assert file_name == "test_data_09_02_2000.csv"


def test_get_filename_xlsx():
    with freeze_time("2000-02-09"):
        file_name = get_filename("test", FileTypes.XLSX)

        assert file_name == "test_data_09_02_2000.xlsx"


def test_get_product_queryset_all(product_list):
    queryset = get_product_queryset({"all": ""})

    assert queryset.count() == len(product_list)


def test_get_product_queryset_ids(product_list):
    pks = [product.pk for product in product_list[:2]]
    queryset = get_product_queryset({"ids": pks})

    assert queryset.count() == len(pks)


def get_product_queryset_filter(product_list):
    product_not_published = product_list.first()
    product_not_published.is_published = False
    product_not_published.save()

    queryset = get_product_queryset({"ids": {"is_published": True}})

    assert queryset.count() == len(product_list) - 1


def test_create_csv_file_and_save_in_export_file_csv(export_file, tmpdir):
    from django.conf import settings

    settings.MEDIA_ROOT = tmpdir

    export_data = [
        {"id": "123", "name": "test1", "collections": "coll1"},
        {"id": "345", "name": "test2"},
    ]
    headers = ["id", "name", "collections"]
    csv_headers_mapping = {"id": "ID", "name": "NAME"}
    delimiter = ";"
    export_file = export_file
    file_name = "test.csv"

    export_file_csv_upload_dir = ExportFile.content_file.field.upload_to

    assert not export_file.content_file

    create_csv_file_and_save_in_export_file(
        export_data,
        headers,
        csv_headers_mapping,
        delimiter,
        export_file,
        file_name,
        FileTypes.CSV,
    )

    csv_file = export_file.content_file
    assert csv_file
    assert csv_file.name == f"{export_file_csv_upload_dir}/{file_name}"

    file_content = csv_file.read().decode().split("\r\n")
    headers = list(csv_headers_mapping.values())
    headers.append("collections")

    assert ";".join(headers) in file_content
    assert ";".join(export_data[0].values()) in file_content
    assert (";".join(export_data[1].values()) + "; ") in file_content

    shutil.rmtree(tmpdir)


def test_create_csv_file_and_save_in_export_file_xlsx(export_file, tmpdir):
    from django.conf import settings

    settings.MEDIA_ROOT = tmpdir

    export_data = [
        {"id": "123", "name": "test1", "collections": "coll1"},
        {"id": "345", "name": "test2"},
    ]
    headers = ["id", "name", "collections"]
    csv_headers_mapping = {"id": "ID", "name": "NAME", "collections": "COLLECTIONS"}
    delimiter = ";"
    export_file = export_file
    file_name = "test.xlsx"

    export_file_csv_upload_dir = ExportFile.content_file.field.upload_to

    assert not export_file.content_file

    create_csv_file_and_save_in_export_file(
        export_data,
        headers,
        csv_headers_mapping,
        delimiter,
        export_file,
        file_name,
        FileTypes.XLSX,
    )

    xlsx_file = export_file.content_file
    assert xlsx_file
    assert xlsx_file.name == f"{export_file_csv_upload_dir}/{file_name}"

    wb_obj = openpyxl.load_workbook(xlsx_file)

    sheet_obj = wb_obj.active
    max_col = sheet_obj.max_column
    max_row = sheet_obj.max_row
    expected_headers = list(csv_headers_mapping.values())
    headers = [sheet_obj.cell(row=1, column=i).value for i in range(1, max_col + 1)]
    data = []
    for i in range(2, max_row + 1):
        row = []
        for j in range(1, max_col + 1):
            row.append(sheet_obj.cell(row=i, column=j).value)
        data.append(row)

    assert headers == expected_headers
    assert list(export_data[0].values()) in data
    row2 = list(export_data[1].values())
    # add string with space for collections column
    row2.append(" ")
    assert row2 in data

    shutil.rmtree(tmpdir)


def test_save_csv_file_in_export_file(export_file, tmpdir):
    from django.conf import settings

    settings.MEDIA_ROOT = tmpdir

    file_mock = MagicMock(spec=File)
    file_mock.name = "temp_file.csv"
    file_name = "test.csv"

    assert not export_file.content_file

    save_csv_file_in_export_file(export_file, file_mock, file_name)

    export_file.refresh_from_db()
    assert export_file.content_file

    shutil.rmtree(tmpdir)